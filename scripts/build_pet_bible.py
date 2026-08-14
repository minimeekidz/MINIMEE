#!/usr/bin/env python3
"""Turn Em's pet spec workbook into JSON the app and other agents can read.

    python3 scripts/build_pet_bible.py

Source:  docs/source/MINIMEE_寵物設定與VO規格_v1.xlsx   (the authority)
Output:  src/data/petBible.json                        (generated, do not edit)

The workbook is the source of truth, not this output and not the code. Em
maintains the spreadsheet; everything downstream is regenerated from it. That
is the whole point of having a converter rather than transcribing the values:
a hand-copied VO line drifts from the sheet the first time anyone edits either
one, and nobody notices until a pet says the wrong thing.

Each sheet has a title row, a note row, then a header row, then data. The
header row is found rather than assumed, so adding a line to a note does not
silently shift every column by one.
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - developer machine only
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "docs/source/MINIMEE_寵物設定與VO規格_v1.xlsx"
OUTPUT = ROOT / "src/data/petBible.json"

# Sheet name -> key in the JSON. Renaming a sheet breaks this on purpose:
# a silently missing section is worse than a loud failure.
SHEETS = {
    "00_玩法總規則": "rules",
    "01_12寵物Master": "pets",
    "02_12級好感度": "levels",
    "03_神秘獎勵": "rewards",
    "04_事件規則庫": "events",
    "05_寵物事件偏好": "petEventPreferences",
    "06_事件VO台詞庫": "eventVo",
    "07_答題VO與狀態": "quizVo",
    "08_出沒地點規則": "spawn",
    "09_ClaudeCode變更": "codeChanges",
    "10_QA測試案例": "qa",
}


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def read_sheet(ws):
    """Rows as dicts, plus the sheet's title and note lines."""
    rows = [[clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]

    # The header is the first row with several filled cells; the lines above it
    # are the human preamble Em writes for each sheet.
    header_index = None
    for index, row in enumerate(rows):
        if sum(1 for cell in row if cell) >= 3:
            header_index = index
            break
    if header_index is None:
        return {"notes": [], "columns": [], "rows": []}

    notes = [row[0] for row in rows[:header_index] if row and row[0]]
    header = rows[header_index]
    columns = [cell for cell in header if cell]

    records = []
    for row in rows[header_index + 1:]:
        if not any(row):
            continue
        record = {}
        for column, cell in zip(header, row):
            if column:
                record[column] = cell
        # A row that only repeats the header, or is entirely blank under it, is
        # spreadsheet padding rather than data.
        if any(record.values()):
            records.append(record)
    return {"notes": notes, "columns": columns, "rows": records}


def main():
    if not SOURCE.exists():
        sys.exit(f"missing source workbook: {SOURCE}")

    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    missing = [name for name in SHEETS if name not in wb.sheetnames]
    if missing:
        sys.exit(f"workbook is missing expected sheets: {missing}")

    bible = {
        "_source": SOURCE.name,
        "_generatedBy": "scripts/build_pet_bible.py",
        "_warning": "Generated from the workbook. Edit the workbook, then rerun the script.",
    }
    for sheet_name, key in SHEETS.items():
        bible[key] = read_sheet(wb[sheet_name])
        print(f"{sheet_name:24} -> {key:22} {len(bible[key]['rows']):3} rows")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(bible, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nwrote {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
