# Converts Em's 36-theme workbook into the JSON the app reads.
#
#   python3 scripts/build_theme_book.py
#
# Same contract as scripts/build_pet_bible.py: the workbook is the source of
# truth, this only reshapes it. Nothing here invents a word, a question or a
# card mapping — if a cell is empty the field is empty, and that shows up in
# the app as "not configured yet" rather than as something plausible.
#
# Re-run it whenever Em updates the workbook.

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

SOURCE = Path("docs/source/MINIMEE_36主題_ActiveTheme_v1.xlsm")
OUT = Path("src/data/themeBook.json")


def header_row(ws, wanted):
    """Finds the row that carries the headers, rather than assuming row 1."""
    for index, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if wanted in cells:
            return index, cells
    raise SystemExit(f"{ws.title}: no header row containing {wanted!r}")


def rows(ws, wanted):
    start, headers = header_row(ws, wanted)
    for row in ws.iter_rows(min_row=start + 1, values_only=True):
        record = {}
        for key, value in zip(headers, row):
            if not key:
                continue
            record[key] = "" if value is None else str(value).strip()
        if any(record.values()):
            yield record


def number(text):
    """'1.0' -> 1. The workbook stores integers as floats."""
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def words(text):
    """'港鐵｜電車｜火車｜輕鐵' -> the four words."""
    if not text:
        return []
    return [part.strip() for part in re.split(r"[｜|/、]", text) if part.strip()]


def main():
    if not SOURCE.exists():
        sys.exit(f"missing {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    for sheet in ("36主題統一版", "MEE卡01-24", "ActiveTheme_v1"):
        if sheet not in wb.sheetnames:
            sys.exit(f"{SOURCE}: missing sheet {sheet!r}")

    themes = []
    for row in rows(wb["36主題統一版"], "Theme ID"):
        theme_id = row.get("Theme ID", "")
        if not theme_id:
            continue
        themes.append({
            "themeId": theme_id,
            "themeNo": number(row.get("主題序")),
            "nameZh": row.get("主題", ""),
            # The unified 3–12 vocabulary. The 3–5 column is kept out of the
            # app on purpose: two vocabularies for one theme is how a child
            # ends up quizzed on a word the video never said.
            "words": words(row.get("統一3-12歲詞彙", "")),
            "vo": row.get("統一粵語VO腳本", ""),
            "question": row.get("VO提問", ""),
            "answerPattern": row.get("回答句式", ""),
        })

    cards = []
    for row in rows(wb["MEE卡01-24"], "Card ID"):
        code = row.get("Card ID", "")
        if not code:
            continue
        cards.append({
            "code": code,
            "cardNumber": number(row.get("Card No.")),
            "bookNo": number(row.get("Book")),
            "slotNo": number(row.get("Slot")),
            "position": row.get("位置", ""),
            "normalAsset": row.get("NORMAL 檔名", ""),
            "flashAsset": row.get("FLASH 檔名", ""),
            "art": row.get("卡面視覺", ""),
            # Em's own note on this sheet: 卡號／Book／Slot 為固定；主題配對屬
            # 建議. So the suggestion is carried, clearly labelled, and the
            # real mapping comes from ActiveTheme_v1.
            "suggestedTheme": row.get("建議對應主題", ""),
            "confidence": row.get("對應信心", ""),
        })

    releases = []
    for row in rows(wb["ActiveTheme_v1"], "Release ID"):
        release_id = row.get("Release ID", "")
        if not release_id:
            continue
        releases.append({
            "releaseId": release_id,
            "traySlot": number(row.get("Tray Slot")),
            "themeId": row.get("Theme ID", ""),
            "nameZh": row.get("主題", ""),
            "words": words(row.get("統一詞彙", "")),
            "status": row.get("狀態", ""),
            "displayOrder": number(row.get("Display Order")),
            "targetCard": row.get("Target Card", ""),
            "bookNo": number(row.get("Book")),
            "slotNo": number(row.get("Slot")),
            "activeFrom": row.get("Active From", "") or None,
            "activeTo": row.get("Active To", "") or None,
        })

    book = {
        "_source": SOURCE.name,
        "_generatedBy": "scripts/build_theme_book.py",
        "_warning": "GENERATED — edit the workbook, not this file.",
        "themes": themes,
        "cards": cards,
        "releases": releases,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(book, ensure_ascii=False, indent=1), encoding="utf-8")

    missing_words = [t["themeId"] for t in themes if len(t["words"]) != 4]
    print(f"themes {len(themes)}  cards {len(cards)}  releases {len(releases)}")
    if missing_words:
        print(f"  ** themes without exactly 4 words: {missing_words}")
    print(f"wrote {OUT}")


main()
